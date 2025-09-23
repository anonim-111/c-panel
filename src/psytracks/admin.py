import datetime
from io import BytesIO

import openpyxl
from django.contrib import admin
from django.contrib.admin import SimpleListFilter
from django.db.models import ExpressionWrapper, F, DateField, DurationField, Case, When, Value, BooleanField
from django.db.models.functions import Greatest, Concat, Cast
from django.http import HttpResponse
from django.urls import path
from django.utils import timezone
from django.utils.html import format_html
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from psytracks.models import SocialDomesticEnvironment, ReasonForSpecialConsideration, Doctor, Patient, Psychiatrist
from psytracks.forms import PatientForm
from django.utils.translation import gettext_lazy as _

from utils.models import Inspector, Neighborhood, District


class PsychiatristFilter(SimpleListFilter):
    title = _("psychiatrist")
    parameter_name = "psychiatrist"

    def lookups(self, request, model_admin):
        qs = model_admin.get_queryset(request)

        user = request.user
        if hasattr(user, "inspector"):
            qs = qs.filter(inspector=user.inspector)
        elif hasattr(user, "neighborhood"):
            qs = qs.filter(neighborhood=user.neighborhood)
        elif hasattr(user, "district"):
            qs = qs.filter(neighborhood__district=user.district.district)
        elif hasattr(user, "region"):
            qs = qs.filter(neighborhood__district__region=user.region.region)

        psychiatrist_ids = qs.values_list("psychiatrist_id", flat=True).distinct()
        psychiatrists = Psychiatrist.objects.filter(id__in=psychiatrist_ids).values_list("id", "full_name")
        return [(d[0], d[1]) for d in psychiatrists if d[0]]

    def queryset(self, request, queryset):
        user = request.user
        if hasattr(user, "inspector"):
            queryset = queryset.filter(inspector=user.inspector)
        elif hasattr(user, "neighborhood"):
            queryset = queryset.filter(neighborhood=user.neighborhood)
        elif hasattr(user, "district"):
            queryset = queryset.filter(neighborhood__district=user.district.district)
        elif hasattr(user, "region"):
            queryset = queryset.filter(neighborhood__district__region=user.region.region)

        if self.value():
            queryset = queryset.filter(psychiatrist_id=self.value())
        return queryset


class InspectorFilter(SimpleListFilter):
    title = _("inspector")
    parameter_name = "inspector"

    def lookups(self, request, model_admin):
        qs = model_admin.get_queryset(request)

        user = request.user
        if hasattr(user, "inspector"):
            qs = qs.filter(inspector=user.inspector)
        elif hasattr(user, "neighborhood"):
            qs = qs.filter(neighborhood=user.neighborhood)
        elif hasattr(user, "district"):
            qs = qs.filter(neighborhood__district=user.district.district)
        elif hasattr(user, "region"):
            qs = qs.filter(neighborhood__district__region=user.region.region)

        inspector_ids = qs.values_list("inspector_id", flat=True).distinct()
        inspectors = Inspector.objects.filter(id__in=inspector_ids).values_list("id", "full_name")
        return [(d[0], d[1]) for d in inspectors if d[0]]

    def queryset(self, request, queryset):
        user = request.user
        if hasattr(user, "inspector"):
            queryset = queryset.filter(inspector=user.inspector)
        elif hasattr(user, "neighborhood"):
            queryset = queryset.filter(neighborhood=user.neighborhood)
        elif hasattr(user, "district"):
            queryset = queryset.filter(neighborhood__district=user.district.district)
        elif hasattr(user, "region"):
            queryset = queryset.filter(neighborhood__district__region=user.region.region)

        if self.value():
            queryset = queryset.filter(inspector_id=self.value())
        return queryset


class NeighborhoodFilter(SimpleListFilter):
    title = _("neighborhood")
    parameter_name = "neighborhood"

    def lookups(self, request, model_admin):
        qs = model_admin.get_queryset(request)

        user = request.user
        if hasattr(user, "inspector"):
            qs = qs.filter(inspector=user.inspector)
        elif hasattr(user, "neighborhood"):
            qs = qs.filter(neighborhood=user.neighborhood)
        elif hasattr(user, "district"):
            qs = qs.filter(neighborhood__district=user.district.district)
        elif hasattr(user, "region"):
            qs = qs.filter(neighborhood__district__region=user.region.region)

        neighborhood_ids = qs.values_list("neighborhood_id", flat=True).distinct()
        neighborhoods = Neighborhood.objects.filter(id__in=neighborhood_ids).annotate(
            str_name=Concat(F("name"), Value(" ("), F("district__name"), Value(")")),
        ).values_list("id", "str_name")
        return [(d[0], d[1]) for d in neighborhoods if d[0]]

    def queryset(self, request, queryset):
        user = request.user
        if hasattr(user, "inspector"):
            queryset = queryset.filter(inspector=user.inspector)
        elif hasattr(user, "neighborhood"):
            queryset = queryset.filter(neighborhood=user.neighborhood)
        elif hasattr(user, "district"):
            queryset = queryset.filter(neighborhood__district=user.district.district)
        elif hasattr(user, "region"):
            queryset = queryset.filter(neighborhood__district__region=user.region.region)

        if self.value():
            queryset = queryset.filter(neighborhood_id=self.value())
        return queryset


class DistrictFilter(SimpleListFilter):
    title = _("district")
    parameter_name = "district"

    def lookups(self, request, model_admin):
        qs = model_admin.get_queryset(request)

        user = request.user
        if hasattr(user, "inspector"):
            qs = qs.filter(inspector=user.inspector)
        elif hasattr(user, "neighborhood"):
            qs = qs.filter(neighborhood=user.neighborhood)
        elif hasattr(user, "district"):
            qs = qs.filter(neighborhood__district=user.district.district)
        elif hasattr(user, "region"):
            qs = qs.filter(neighborhood__district__region=user.region.region)

        district_ids = qs.values_list("neighborhood__district_id", flat=True).distinct()
        districts = District.objects.filter(id__in=district_ids).values_list("id", "name")
        return [(d[0], d[1]) for d in districts if d[0]]

    def queryset(self, request, queryset):
        user = request.user
        if hasattr(user, "inspector"):
            queryset = queryset.filter(inspector=user.inspector)
        elif hasattr(user, "neighborhood"):
            queryset = queryset.filter(neighborhood=user.neighborhood)
        elif hasattr(user, "district"):
            queryset = queryset.filter(neighborhood__district=user.district.district)
        elif hasattr(user, "region"):
            queryset = queryset.filter(neighborhood__district__region=user.region.region)

        if self.value():
            queryset = queryset.filter(neighborhood__district_id=self.value())
        return queryset


class DistrictDoctorFilter(SimpleListFilter):
    title = _("district")
    parameter_name = "district"

    def lookups(self, request, model_admin):
        qs = model_admin.get_queryset(request)

        user = request.user
        if hasattr(user, "inspector"):
            qs = qs.filter(neighborhood__inspector=user.inspector)
        elif hasattr(user, "neighborhood"):
            qs = qs.filter(neighborhood=user.neighborhood)
        elif hasattr(user, "district"):
            qs = qs.filter(neighborhood__district=user.district.district)
        elif hasattr(user, "region"):
            qs = qs.filter(neighborhood__district__region=user.region.region)

        district_ids = qs.values_list("neighborhood__district_id", flat=True).distinct()
        districts = District.objects.filter(id__in=district_ids).values_list("id", "name")
        return [(d[0], d[1]) for d in districts if d[0]]

    def queryset(self, request, queryset):
        user = request.user
        if hasattr(user, "inspector"):
            queryset = queryset.filter(neighborhood__inspector=user.inspector)
        elif hasattr(user, "neighborhood"):
            queryset = queryset.filter(neighborhood=user.neighborhood)
        elif hasattr(user, "district"):
            queryset = queryset.filter(neighborhood__district=user.district.district)
        elif hasattr(user, "region"):
            queryset = queryset.filter(neighborhood__district__region=user.region.region)

        if self.value():
            queryset = queryset.filter(neighborhood__district_id=self.value())
        return queryset


class DistrictPsychiatristFilter(SimpleListFilter):
    title = _("district")
    parameter_name = "district"

    def lookups(self, request, model_admin):
        qs = model_admin.get_queryset(request)

        user = request.user
        if hasattr(user, "neighborhood"):
            qs = qs.filter(district=user.neighborhood.district)
        elif hasattr(user, "district"):
            qs = qs.filter(district=user.district.district)
        elif hasattr(user, "region"):
            qs = qs.filter(district__region=user.region.region)

        district_ids = qs.values_list("district_id", flat=True).distinct()
        districts = District.objects.filter(id__in=district_ids).values_list("id", "name")
        return [(d[0], d[1]) for d in districts if d[0]]

    def queryset(self, request, queryset):
        user = request.user
        if hasattr(user, "neighborhood"):
            queryset = queryset.filter(district=user.neighborhood.district)
        elif hasattr(user, "district"):
            queryset = queryset.filter(district=user.district.district)
        elif hasattr(user, "region"):
            queryset = queryset.filter(district__region=user.region.region)

        if self.value():
            queryset = queryset.filter(district_id=self.value())
        return queryset


class NeighborhoodDoctorFilter(SimpleListFilter):
    title = _("neighborhood")
    parameter_name = "neighborhood"

    def lookups(self, request, model_admin):
        qs = model_admin.get_queryset(request)

        user = request.user
        if hasattr(user, "inspector"):
            qs = qs.filter(neighborhood__inspector=user.inspector)
        elif hasattr(user, "neighborhood"):
            qs = qs.filter(neighborhood=user.neighborhood)
        elif hasattr(user, "district"):
            qs = qs.filter(neighborhood__district=user.district.district)
        elif hasattr(user, "region"):
            qs = qs.filter(neighborhood__district__region=user.region.region)

        neighborhood_ids = qs.values_list("neighborhood_id", flat=True).distinct()
        neighborhoods = Neighborhood.objects.filter(id__in=neighborhood_ids).annotate(
            str_name=Concat(F("name"), Value(" ("), F("district__name"), Value(")")),
        ).values_list("id", "str_name")
        return [(d[0], d[1]) for d in neighborhoods if d[0]]

    def queryset(self, request, queryset):
        user = request.user
        if hasattr(user, "inspector"):
            queryset = queryset.filter(neighborhood__inspector=user.inspector)
        elif hasattr(user, "neighborhood"):
            queryset = queryset.filter(neighborhood=user.neighborhood)
        elif hasattr(user, "district"):
            queryset = queryset.filter(neighborhood__district=user.district.district)
        elif hasattr(user, "region"):
            queryset = queryset.filter(neighborhood__district__region=user.region.region)

        if self.value():
            queryset = queryset.filter(neighborhood_id=self.value())
        return queryset


class OverdueFilter(admin.SimpleListFilter):
    title = _("Is overdue")
    parameter_name = "is_overdue"

    def lookups(self, request, model_admin):
        return [
            ("yes", _("Yes")),
            ("no", _("No")),
        ]

    def queryset(self, request, queryset):
        if self.value() == "yes":
            return queryset.filter(is_overdue=True)
        elif self.value() == "no":
            return queryset.filter(is_overdue=False)
        return queryset


@admin.register(SocialDomesticEnvironment)
class SocialDomesticEnvironmentAdmin(admin.ModelAdmin):
    list_display = ("id", "name")
    list_display_links = ("id", "name")


@admin.register(ReasonForSpecialConsideration)
class ReasonForSpecialConsiderationAdmin(admin.ModelAdmin):
    list_display = ("id", "name")
    list_display_links = ("id", "name")


@admin.register(Psychiatrist)
class PsychiatristAdmin(admin.ModelAdmin):
    list_display = ("id", "full_name", "district", "phone", "user")
    list_display_links = ("id", "full_name")
    search_fields = ("full_name__icontains",)
    list_filter = (DistrictPsychiatristFilter,)


@admin.register(Doctor)
class DoctorAdmin(admin.ModelAdmin):
    list_display = ("id", "full_name", "neighborhood", "phone", "fbirth_date", "brigade_number", "polyclinic_name")
    list_display_links = ("id", "full_name")
    search_fields = ("full_name__icontains",)
    list_filter = (NeighborhoodDoctorFilter, DistrictDoctorFilter)

    def fbirth_date(self, obj):
        if obj.birth_date:
            return obj.birth_date.strftime("%d.%m.%Y")
        return "—"

    fbirth_date.admin_order_field = "birth_date"
    fbirth_date.short_description = Doctor._meta.get_field("birth_date").verbose_name

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        user = request.user
        if db_field.name == "neighborhood":
            if hasattr(user, "district"):
                kwargs["queryset"] = Neighborhood.objects.filter(district_id=user.district.district_id)
            elif hasattr(user, 'region'):
                kwargs["queryset"] = Neighborhood.objects.filter(district__region_id=user.region.region_id)
            else:
                kwargs["queryset"] = Neighborhood.objects.all()
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


@admin.register(Patient)
class PatientAdmin(admin.ModelAdmin):
    list_display = ("id", "full_name", "pinfl", "fbirth_date", "is_aggressive", "is_convicted", "is_abroad_long_term",
                    "neighborhood__district", "neighborhood__name", "address", "inspector", "psychiatrist",
                    "flast_psychiatric_appointment_date", "flast_home_visit_by_doctor_date", "reason",
                    "receiving_supportive_therapy", "reason_for_special_consideration", "description_for_special_consideration",
                    "alcohol_and_drug_use", "where_is_now", "description_where_is_now", "flast_hospitalization_from",
                    "flast_hospitalization_to", "next_psychiatric_appointment_date", "last_psychiatric_appointment_days_left"  )
    list_display_links = ("id", "full_name")
    form = PatientForm
    autocomplete_fields = ("psychiatrist", "inspector")
    search_fields = ["full_name__icontains"]
    change_list_template = "admin/patients_changelist.html"
    list_filter = [DistrictFilter, NeighborhoodFilter, PsychiatristFilter, InspectorFilter, "is_aggressive",
                   "is_convicted", "is_abroad_long_term", OverdueFilter]

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        user = request.user
        if db_field.name == "neighborhood":
            if hasattr(user, "district"):
                kwargs["queryset"] = Neighborhood.objects.filter(district_id=user.district.district_id)
            elif hasattr(user, 'region'):
                kwargs["queryset"] = Neighborhood.objects.filter(district__region_id=user.region.region_id)
            else:
                kwargs["queryset"] = Neighborhood.objects.all()
        if db_field.name == "inspector":
            if hasattr(user, "district"):
                kwargs["queryset"] = Inspector.objects.filter(neighborhood__district_id=user.district.district_id)
            elif hasattr(user, 'region'):
                kwargs["queryset"] = Inspector.objects.filter(neighborhood__district__region_id=user.region.region_id)
            else:
                kwargs["queryset"] = Inspector.objects.all()
        if db_field.name == "psychiatrist":
            if hasattr(user, "district"):
                kwargs["queryset"] = Psychiatrist.objects.filter(district_id=user.district.district_id)
            elif hasattr(user, 'region'):
                kwargs["queryset"] = Psychiatrist.objects.filter(district__region_id=user.region.region_id)
            else:
                kwargs["queryset"] = Psychiatrist.objects.all()
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def fbirth_date(self, obj):
        if obj.birth_date:
            return obj.birth_date.strftime("%d.%m.%Y")
        return "—"

    fbirth_date.admin_order_field = "birth_date"  # ordering ishlaydi
    fbirth_date.short_description = Patient._meta.get_field("birth_date").verbose_name

    def flast_psychiatric_appointment_date(self, obj):
        if obj.last_psychiatric_appointment_date:
            return obj.last_psychiatric_appointment_date.strftime("%d.%m.%Y")
        return "—"

    flast_psychiatric_appointment_date.admin_order_field = "last_psychiatric_appointment_date"  # ordering ishlaydi
    flast_psychiatric_appointment_date.short_description = Patient._meta.get_field("last_psychiatric_appointment_date").verbose_name

    def flast_home_visit_by_doctor_date(self, obj):
        if obj.last_home_visit_by_doctor_date:
            return obj.last_home_visit_by_doctor_date.strftime("%d.%m.%Y")
        return "—"

    flast_home_visit_by_doctor_date.admin_order_field = "last_home_visit_by_doctor_date"  # ordering ishlaydi
    flast_home_visit_by_doctor_date.short_description = Patient._meta.get_field("last_home_visit_by_doctor_date").verbose_name

    def flast_hospitalization_from(self, obj):
        if obj.last_hospitalization_from:
            return obj.last_hospitalization_from.strftime("%d.%m.%Y")
        return "—"

    flast_hospitalization_from.admin_order_field = "last_hospitalization_from"  # ordering ishlaydi
    flast_hospitalization_from.short_description = Patient._meta.get_field("last_hospitalization_from").verbose_name

    def flast_hospitalization_to(self, obj):
        if obj.last_hospitalization_to:
            return obj.last_hospitalization_to.strftime("%d.%m.%Y")
        return "—"

    flast_hospitalization_to.admin_order_field = "last_hospitalization_to"  # ordering ishlaydi
    flast_hospitalization_to.short_description = Patient._meta.get_field("last_hospitalization_to").verbose_name

    def get_readonly_fields(self, request, obj=None):
        user = request.user
        if hasattr(user, "inspector"):
            return ["full_name", "pinfl", "fbirth_date", "is_aggressive", "neighborhood_display", "address", "inspector_display",
                    "psychiatrist_display", "reason_for_special_consideration_display", "description_for_special_consideration",
                    "reason", "flast_psychiatric_appointment_date", "flast_home_visit_by_doctor_date",
                     "flast_hospitalization_from", "receiving_supportive_therapy", "alcohol_and_drug_use",
                    "flast_hospitalization_to", "last_psychiatric_appointment_file", "last_home_visit_by_doctor_file",
                    "last_hospitalization_from_file", "last_hospitalization_to_file", "max_examination_interval"]
        elif hasattr(user, "neighborhood"):
            return ["full_name", "pinfl", "fbirth_date", "is_aggressive", "is_convicted", "is_abroad_long_term",
                    "neighborhood_display", "address", "inspector_display", "psychiatrist_display",
                    "reason_for_special_consideration_display", "description_for_special_consideration", "where_is_now",
                    "description_where_is_now", "social_domestic_environment_display", "last_psychiatric_appointment_file", "flast_psychiatric_appointment_date",
                    "flast_home_visit_by_doctor_date", "flast_hospitalization_from", "flast_hospitalization_to", "max_examination_interval"]
        elif hasattr(user, "psychiatrist"):
            return ["full_name", "pinfl", "fbirth_date", "is_aggressive", "is_convicted", "is_abroad_long_term",
                    "neighborhood_display", "address", "inspector_display", "psychiatrist_display",
                    "reason_for_special_consideration_display", "description_for_special_consideration", "where_is_now",
                    "description_where_is_now", "social_domestic_environment_display", "flast_psychiatric_appointment_date",
                    "flast_home_visit_by_doctor_date", "flast_hospitalization_from", "flast_hospitalization_to", "max_examination_interval",
                    "last_home_visit_by_doctor_file", "last_hospitalization_from_file", "last_hospitalization_to_file",
                    "receiving_supportive_therapy", "alcohol_and_drug_use", "reason"]
        return []

    def get_fields(self, request, obj=None):
        fields = list(super().get_fields(request, obj))
        user = request.user

        if hasattr(user, "neighborhood"):
            if "neighborhood" in fields:
                fields.remove("neighborhood")
            if "psychiatrist" in fields:
                fields.remove("psychiatrist")
            if "inspector" in fields:
                fields.remove("inspector")
            if "social_domestic_environment" in fields:
                fields.remove("social_domestic_environment")
            if "reason_for_special_consideration" in fields:
                fields.remove("reason_for_special_consideration")
            if "birth_date" in fields:
                fields.remove("birth_date")
            if "last_psychiatric_appointment_date" in fields:
                fields.remove("last_psychiatric_appointment_date")
            if "last_home_visit_by_doctor_date" in fields:
                fields.remove("last_home_visit_by_doctor_date")
            if "last_hospitalization_from" in fields:
                fields.remove("last_hospitalization_from")
            if "last_hospitalization_to" in fields:
                fields.remove("last_hospitalization_to")
        elif hasattr(user, "psychiatrist"):
            if "neighborhood" in fields:
                fields.remove("neighborhood")
            if "psychiatrist" in fields:
                fields.remove("psychiatrist")
            if "inspector" in fields:
                fields.remove("inspector")
            if "social_domestic_environment" in fields:
                fields.remove("social_domestic_environment")
            if "reason_for_special_consideration" in fields:
                fields.remove("reason_for_special_consideration")
            if "birth_date" in fields:
                fields.remove("birth_date")
            if "last_psychiatric_appointment_date" in fields:
                fields.remove("last_psychiatric_appointment_date")
            if "last_home_visit_by_doctor_date" in fields:
                fields.remove("last_home_visit_by_doctor_date")
            if "last_hospitalization_from" in fields:
                fields.remove("last_hospitalization_from")
            if "last_hospitalization_to" in fields:
                fields.remove("last_hospitalization_to")
        elif hasattr(user, "inspector"):
            if "neighborhood" in fields:
                fields.remove("neighborhood")
            if "psychiatrist" in fields:
                fields.remove("psychiatrist")
            if "inspector" in fields:
                fields.remove("inspector")
            if "reason_for_special_consideration" in fields:
                fields.remove("reason_for_special_consideration")
            if "last_psychiatric_appointment_file" in fields:
                fields.remove("last_psychiatric_appointment_file")
            if "last_home_visit_by_doctor_file" in fields:
                fields.remove("last_home_visit_by_doctor_file")
            if "last_hospitalization_from_file" in fields:
                fields.remove("last_hospitalization_from_file")
            if "last_hospitalization_to_file" in fields:
                fields.remove("last_hospitalization_to_file")
            if "birth_date" in fields:
                fields.remove("birth_date")
            if "last_psychiatric_appointment_date" in fields:
                fields.remove("last_psychiatric_appointment_date")
            if "last_home_visit_by_doctor_date" in fields:
                fields.remove("last_home_visit_by_doctor_date")
            if "last_hospitalization_from" in fields:
                fields.remove("last_hospitalization_from")
            if "last_hospitalization_to" in fields:
                fields.remove("last_hospitalization_to")
        return fields

    def get_queryset(self, request):
        today = timezone.now().date()
        queryset = super(PatientAdmin, self).get_queryset(request)
        user = request.user
        if hasattr(user, "inspector"):
            queryset = queryset.filter(inspector__user=user)
        elif hasattr(user, "neighborhood"):
            queryset = queryset.filter(neighborhood__user=user)
        elif hasattr(user, "psychiatrist"):
            queryset = queryset.filter(psychiatrist__user=user)
        elif hasattr(user, "district"):
            queryset = queryset.filter(neighborhood__district=user.district.district)
        elif hasattr(user, "region"):
            queryset = queryset.filter(neighborhood__district__region=user.region.region)
        return queryset.annotate(
            last_date=Case(
                When(
                    last_hospitalization_from__isnull=False,
                    last_hospitalization_to__isnull=False,
                    last_hospitalization_from__gt=F("last_hospitalization_to"),
                    then=Value(today)
                ),
                When(
                    last_hospitalization_from__isnull=False,
                    last_hospitalization_to__isnull=True,
                    then=Value(today)
                ),
                When(
                    last_hospitalization_from__isnull=True,
                    last_hospitalization_to__isnull=False,
                    then=Greatest(
                        "last_hospitalization_to",
                        "last_psychiatric_appointment_date",
                        "last_home_visit_by_doctor_date",
                    )
                ),
                When(
                    last_hospitalization_from__isnull=False,
                    last_hospitalization_to__isnull=False,
                    last_hospitalization_from__lte=F("last_hospitalization_to"),
                    then=Greatest(
                        "last_hospitalization_to",
                        "last_psychiatric_appointment_date",
                        "last_home_visit_by_doctor_date",
                    )
                ),
                When(
                    last_hospitalization_from__isnull=True,
                    last_hospitalization_to__isnull=True,
                    then=Greatest(
                        "last_psychiatric_appointment_date",
                        "last_home_visit_by_doctor_date",
                    )
                ),
                default=None,
                output_field=DateField()
            )
        ).annotate(
            interval=ExpressionWrapper(
                F("max_examination_interval") * Value(datetime.timedelta(days=1)),
                output_field=DurationField()
            ),
            deadline=Cast(
                ExpressionWrapper(
                    F("last_date") + F("interval"),
                    output_field=DateField(),
                ),
                DateField()
            ),
            is_overdue=Case(
                When(last_date__isnull=True, then=Value(True)),
                When(deadline__lt=Value(today, output_field=DateField()), then=Value(True)),
                default=Value(False),
                output_field=BooleanField()
            )
        )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("export-excel/", self.admin_site.admin_view(self.export_as_excel), name="patients_export_excel"),
        ]
        return custom_urls + urls

    def neighborhood__district(self, obj):
        return obj.neighborhood.district.name

    neighborhood__district.short_description = _("district")

    def neighborhood__name(self, obj):
        return obj.neighborhood.name

    neighborhood__name.short_description = _("neighborhood")

    def neighborhood_display(self, obj):
        if obj and obj.neighborhood:
            return format_html("{}", obj.neighborhood)
        return "-"

    neighborhood_display.short_description = _("neighborhood")

    def psychiatrist_display(self, obj):
        if obj and obj.psychiatrist:
            return format_html("{}", obj.psychiatrist)
        return "-"

    psychiatrist_display.short_description = _("psychiatrist")

    def reason_for_special_consideration_display(self, obj):
        if obj and obj.reason_for_special_consideration:
            return format_html("{}", obj.reason_for_special_consideration)
        return "-"

    reason_for_special_consideration_display.short_description = _("reason for special consideration")

    def social_domestic_environment_display(self, obj):
        if obj and obj.social_domestic_environment:
            return format_html("{}", obj.social_domestic_environment)
        return "-"

    social_domestic_environment_display.short_description = _("social-domestic environment")

    def inspector_display(self, obj):
        if obj and obj.inspector:
            return format_html("{}", obj.inspector)
        return "-"

    inspector_display.short_description = _("inspector")

    @admin.display(description=_("next medical examination deadline"),
                   ordering="deadline")
    def next_psychiatric_appointment_date(self, obj):
        if obj.deadline:
            return obj.deadline.strftime("%d.%m.%Y")
        return "—"

    @admin.display(description=_("days left until the next medical examination"),
                   ordering="deadline")
    def last_psychiatric_appointment_days_left(self, obj):
        if obj.deadline:
            today = timezone.now().date()
            days = (obj.deadline - today).days
            if days < 3:
                color = "#f8d7da"
            elif days < 7:
                color = "#fff3cd"
            else:
                color = "#d4edda"
            if days < 0:
                text = _("%(days)d days past") % {"days": abs(days)}
            else:
                text = _("%(days)d days left") % {"days": days}

            return format_html(
                '<div style="background-color:{}; padding:4px; text-align:center;">{}</div>',
                color, text
            )

        else:
            return None

    def export_as_excel(self, request):
        qs = self.get_changelist_instance(request).get_queryset(request)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ruhiy kasallar"

        ws.append(["№", "Ф.И.Ш.", "ПИНФЛ", "Туг.йили", "Ҳудуд", "Маҳалла", "Манзили", "Махсус хисобга олиш сабаби",
                   "Махсус хисобга олишга изоҳи",
                   "Охирги психиатр кабулига келган куни", "Охирги шифокор ёки хамшира томонидан уйдаги курик",
                   "Агар бир ой ичида кўрилмаган бўлса сабабини қўрсатилсин", "Кувватловчи терапия олиниши",
                   "Ижтимоий-Маиший муҳит", "Алкоголь, наркотик моддалар истемол килиши", "Охирги госпитализация",
                   "Бемор ҳозирги кунда каерда", "Бемор ҳозирги кунда каерда изоҳ",	"Худуд еки Туман психиатр Ф.И.Ш.",
                   "Бириктирилган Ички ишлар ходими", "Аҳоли учун хавф туғдираяптими", "Муқаддам судланганми",
                   "Узоқ муддатга кетганми"])

        for index, obj in enumerate(qs, start=1):
            ws.append([
                index,
                obj.full_name,
                obj.pinfl,
                obj.birth_date.strftime("%d.%m.%Y") if obj.birth_date else "",
                obj.neighborhood.district.name if obj.neighborhood and obj.neighborhood.district else "",
                obj.neighborhood.name if obj.neighborhood else "",
                obj.address,
                obj.reason_for_special_consideration.name if obj.reason_for_special_consideration else "",
                obj.description_for_special_consideration,
                obj.last_psychiatric_appointment_date.strftime("%d.%m.%Y") if obj.last_psychiatric_appointment_date else "",
                obj.last_home_visit_by_doctor_date.strftime("%d.%m.%Y") if obj.last_home_visit_by_doctor_date else "",
                obj.reason,
                obj.get_receiving_supportive_therapy_display(),
                obj.social_domestic_environment.name if obj.social_domestic_environment else "",
                obj.get_alcohol_and_drug_use_display(),
                (
                    obj.last_hospitalization_from.strftime("%d.%m.%Y") + "-" + (
                        "хозиргача"
                        if (obj.last_hospitalization_from and (
                                not obj.last_hospitalization_to or obj.last_hospitalization_from > obj.last_hospitalization_to
                        ))
                        else obj.last_hospitalization_to.strftime("%d.%m.%Y")
                    )
                    if obj and obj.last_hospitalization_from
                    else "номаълум санадан - " + obj.last_hospitalization_to.strftime("%d.%m.%Y")
                    if obj and obj.last_hospitalization_to
                    else ""
                ),
                obj.get_where_is_now_display(),
                obj.description_where_is_now,
                obj.psychiatrist.full_name if obj.psychiatrist else "",
                obj.inspector.full_name if obj.inspector else "",
                "Ҳа" if obj.is_aggressive else "Йўқ",
                "Ҳа" if obj.is_convicted else "Йўқ",
                "Ҳа" if obj.is_abroad_long_term else "Йўқ",
            ])

        uniform_width = 20
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = uniform_width
            for cell in col:
                cell.alignment = Alignment(wrap_text=True)

        today_str = datetime.date.today().strftime("%Y-%m-%d")
        filename = f"patients-{today_str}.xlsx"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename={filename}'
        return response