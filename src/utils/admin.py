import datetime
from collections import defaultdict
from datetime import timedelta
from io import BytesIO

import openpyxl
from django.contrib import admin
from django.contrib.admin import SimpleListFilter
from django.contrib.admin.utils import unquote
from django.db.models import Count, Q, Case, When, Value, F, DateField, ExpressionWrapper, DurationField, BooleanField, \
    Sum
from django.db.models.functions import Greatest, Concat
from django.http import HttpResponse
from django.template.response import TemplateResponse
from django.urls import path
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from psytracks.models import Patient
from utils.models import Region, District, Neighborhood, Inspector, SettingsKey, DistrictMonitoring


class DistrictNeighborhoodFilter(SimpleListFilter):
    title = _("district")
    parameter_name = "district"

    def lookups(self, request, model_admin):
        qs = model_admin.get_queryset(request)

        user = request.user
        if hasattr(user, "inspector"):
            qs = qs.filter(inspector=user.inspector)
        elif hasattr(user, "neighborhood"):
            qs = qs.filter(user=user)
        elif hasattr(user, "district"):
            qs = qs.filter(district=user.district.district)
        elif hasattr(user, "region"):
            qs = qs.filter(district__region=user.region.region)

        district_ids = qs.values_list("district_id", flat=True).distinct()
        districts = District.objects.filter(id__in=district_ids).values_list("id", "name")
        return [(d[0], d[1]) for d in districts if d[0]]

    def queryset(self, request, queryset):
        user = request.user
        if hasattr(user, "inspector"):
            queryset = queryset.filter(inspector=user.inspector)
        elif hasattr(user, "neighborhood"):
            queryset = queryset.filter(user=user)
        elif hasattr(user, "district"):
            queryset = queryset.filter(district=user.district.district)
        elif hasattr(user, "region"):
            queryset = queryset.filter(district__region=user.region.region)

        if self.value():
            queryset = queryset.filter(district_id=self.value())
        return queryset


class DistrictInspectorFilter(SimpleListFilter):
    title = _("district")
    parameter_name = "district"

    def lookups(self, request, model_admin):
        qs = model_admin.get_queryset(request)

        user = request.user
        if hasattr(user, "inspector"):
            qs = qs.filter(user=user)
        elif hasattr(user, "neighborhood"):
            qs = qs.filter(neighborhood=user.neighborhood)
        elif hasattr(user, "district"):
            qs = qs.filter(neighborhood__district=user.district.district)
        elif hasattr(user, "region"):
            qs = qs.filter(neighborhood__district__region=user.region.region)

        district_ids = qs.values_list("neighborhood__district_id", flat=True).distinct()
        districts = District.objects.filter(id__in=district_ids).values_list("id", "name")
        return [(d[0], d[1]) for d in set(districts) if d[0]]

    def queryset(self, request, queryset):
        user = request.user
        if hasattr(user, "inspector"):
            queryset = queryset.filter(user=user)
        elif hasattr(user, "neighborhood"):
            queryset = queryset.filter(neighborhood=user.neighborhood)
        elif hasattr(user, "district"):
            queryset = queryset.filter(neighborhood__district=user.district.district)
        elif hasattr(user, "region"):
            queryset = queryset.filter(neighborhood__district__region=user.region.region)

        if self.value():
            queryset = queryset.filter(neighborhood__district_id=self.value())
        return queryset


class NeighborhoodInspectorFilter(SimpleListFilter):
    title = _("neighborhood")
    parameter_name = "neighborhood"

    def lookups(self, request, model_admin):
        qs = model_admin.get_queryset(request)

        user = request.user
        if hasattr(user, "inspector"):
            qs = qs.filter(user=user)
        elif hasattr(user, "neighborhood"):
            qs = qs.filter(neighborhood=user.neighborhood)
        elif hasattr(user, "district"):
            qs = qs.filter(neighborhood__district=user.district.district)
        elif hasattr(user, "region"):
            qs = qs.filter(neighborhood__district__region=user.region.region)

        neighborhood_ids = qs.values_list("neighborhood_id", flat=True).distinct()
        neighborhoods = Neighborhood.objects.filter(id__in=neighborhood_ids).annotate(
            str_name=Concat(F("name"), Value(" ("), F("district__name"), Value(")")),
        ).values_list("id", "str_name")
        return [(d[0], d[1]) for d in neighborhoods if d[0]]

    def queryset(self, request, queryset):
        user = request.user
        if hasattr(user, "inspector"):
            queryset = queryset.filter(user=user)
        elif hasattr(user, "neighborhood"):
            queryset = queryset.filter(neighborhood=user.neighborhood)
        elif hasattr(user, "district"):
            queryset = queryset.filter(neighborhood__district=user.district.district)
        elif hasattr(user, "region"):
            queryset = queryset.filter(neighborhood__district__region=user.region.region)

        if self.value():
            queryset = queryset.filter(neighborhood_id=self.value())
        return queryset


@admin.register(Region)
class RegionAdmin(admin.ModelAdmin):
    list_display = ('id', 'name')
    list_display_links = ('id', 'name')
    search_fields = ["name__icontains"]
    ordering = ('name',)


@admin.register(District)
class DistrictAdmin(admin.ModelAdmin):
    list_display = ('id', 'name', 'region')
    list_display_links = ('id', 'name')
    search_fields = ["name__icontains"]
    ordering = ('name',)
    list_filter = ["region",]


@admin.register(Neighborhood)
class NeighborhoodAdmin(admin.ModelAdmin):
    list_display = ('id', 'name', 'user', 'district')
    list_display_links = ('id', 'name')
    search_fields = ["name__icontains", "district__name__icontains"]
    ordering = ('name',)
    list_filter = ["district__region", DistrictNeighborhoodFilter]
    autocomplete_fields = ["user",]


@admin.register(Inspector)
class InspectorAdmin(admin.ModelAdmin):
    list_display = ('id', 'full_name', 'neighborhood', 'phone', 'user')
    list_display_links = ('id', 'full_name')
    ordering = ('full_name',)
    autocomplete_fields = ["neighborhood", "user"]
    search_fields = ["full_name__icontains"]
    list_filter = ["neighborhood__district__region", DistrictInspectorFilter, NeighborhoodInspectorFilter]


@admin.register(SettingsKey)
class SettingsKeyAdmin(admin.ModelAdmin):
    list_display = ('id', 'name', 'key', 'value')
    list_display_links = ('id', 'name')
    search_fields = ["name__icontains"]



@admin.register(DistrictMonitoring)
class DistrictMonitoringAdmin(admin.ModelAdmin):
    change_list_template = "admin/district_monitoring.html"
    list_display = ("name", "total_neighborhood", "total_patients", "total_aggressive_patients",
                    "total_convicted_patients", "total_abroad_long_term_patients", "late_count",
                "on_time_count", "aggressive_late_count", "aggressive_on_time_count")

    def change_view(self, request, object_id, form_url='', extra_context=None):
        obj = self.get_object(request, unquote(object_id))
        user = request.user
        filter_q = Q(district_id=object_id)
        if hasattr(user, "district"):
            filter_q &= Q(district_id=user.district.district_id)
        elif hasattr(user, "region"):
            filter_q &= Q(district__region_id=user.region.region_id)

        today = timezone.now().date()
        ordering = request.GET.get("o", "name")
        if ordering:
            field = ordering.lstrip("-")
            if field not in self.list_display:
                ordering = field.split("name")
        patients = Patient.objects.annotate(
            last_date=Case(
                When(
                    last_hospitalization_from__isnull=False,
                    last_hospitalization_to__isnull=False,
                    last_hospitalization_from__gt=F("last_hospitalization_to"),
                    then=Value(today)
                ),
                When(
                    last_hospitalization_from__isnull=False,
                    last_hospitalization_to__isnull=True,
                    then=Value(today)
                ),
                When(
                    last_hospitalization_from__isnull=True,
                    last_hospitalization_to__isnull=False,
                    then=Greatest(
                        "last_hospitalization_to",
                        "last_psychiatric_appointment_date",
                        "last_home_visit_by_doctor_date",
                    )
                ),
                When(
                    last_hospitalization_from__isnull=False,
                    last_hospitalization_to__isnull=False,
                    last_hospitalization_from__lte=F("last_hospitalization_to"),
                    then=Greatest(
                        "last_hospitalization_to",
                        "last_psychiatric_appointment_date",
                        "last_home_visit_by_doctor_date",
                    )
                ),
                When(
                    last_hospitalization_from__isnull=True,
                    last_hospitalization_to__isnull=True,
                    then=Greatest(
                        "last_psychiatric_appointment_date",
                        "last_home_visit_by_doctor_date",
                    )
                ),
                default=None,
                output_field=DateField()
            )
        ).annotate(
            interval=ExpressionWrapper(
                F("max_examination_interval") * Value(timedelta(days=1)),
                output_field=DurationField()
            ),
            deadline=ExpressionWrapper(
                F("last_date") + F("interval"),
                output_field=DateField()
            ),
            # deadline o‘tib ketganmi
            is_overdue=Case(
                When(last_date__isnull=True, then=Value(True)),
                When(deadline__lt=Value(today, output_field=DateField()), then=Value(True)),
                default=Value(False),
                output_field=BooleanField()
            )
        )

        qs = Neighborhood.objects.filter(filter_q).annotate(
            total_patients=Count("patients", distinct=True),
            total_aggressive_patients=Count(
                "patients",
                filter=Q(patients__in=patients.filter(is_aggressive=True)),
                distinct=True
            ),
            total_convicted_patients=Count(
                "patients",
                filter=Q(patients__in=patients.filter(is_convicted=True)),
                distinct=True
            ),
            total_abroad_long_term_patients=Count(
                "patients",
                filter=Q(patients__in=patients.filter(is_abroad_long_term=True)),
                distinct=True
            ),
            late_count=Count(
                "patients",
                filter=Q(patients__in=patients.filter(is_overdue=True)),
                distinct=True,
            ),
            on_time_count=Count(
                "patients",
                filter=Q(patients__in=patients.filter(is_overdue=False)),
                distinct=True,
            ),
            aggressive_late_count=Count(
                "patients",
                filter=Q(patients__in=patients.filter(is_aggressive=True, is_overdue=True)),
                distinct=True,
            ),
            aggressive_on_time_count=Count(
                "patients",
                filter=Q(patients__in=patients.filter(is_aggressive=True, is_overdue=False)),
                distinct=True,
            ),
        ).order_by(ordering)

        totals = defaultdict(int)
        for row in qs:
            totals["total_patients"] += row.total_patients
            totals["total_aggressive_patients"] += row.total_aggressive_patients
            totals["total_convicted_patients"] += row.total_convicted_patients
            totals["total_abroad_long_term_patients"] += row.total_abroad_long_term_patients
            totals["on_time_count"] += row.on_time_count
            totals["late_count"] += row.late_count
            totals["aggressive_on_time_count"] += row.aggressive_on_time_count
            totals["aggressive_late_count"] += row.aggressive_late_count

        if "export" in request.GET:
            resp = self.detail_export_as_excel(obj.name, qs, totals)
            return resp

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "original": obj,
            "title": f"{obj} мониторинги",
        }
        if extra_context:
            context.update(extra_context)

        context["neighborhoods"] = qs
        context["totals"] = dict(totals)
        return TemplateResponse(request, "admin/district_monitoring_change_table.html", context)

    def get_queryset(self, request):
        today = timezone.now().date()
        user = request.user
        filter_q = Q(id__gt=0)
        if hasattr(user, "district"):
            filter_q = Q(id=user.district.district_id)
        elif hasattr(user, "region"):
            filter_q = Q(region_id=user.region.region_id)


        patients = Patient.objects.annotate(
            last_date=Case(
                When(
                    last_hospitalization_from__isnull=False,
                    last_hospitalization_to__isnull=False,
                    last_hospitalization_from__gt=F("last_hospitalization_to"),
                    then=Value(today)
                ),
                When(
                    last_hospitalization_from__isnull=False,
                    last_hospitalization_to__isnull=True,
                    then=Value(today)
                ),
                When(
                    last_hospitalization_from__isnull=True,
                    last_hospitalization_to__isnull=False,
                    then=Greatest(
                        "last_hospitalization_to",
                        "last_psychiatric_appointment_date",
                        "last_home_visit_by_doctor_date",
                    )
                ),
                When(
                    last_hospitalization_from__isnull=False,
                    last_hospitalization_to__isnull=False,
                    last_hospitalization_from__lte=F("last_hospitalization_to"),
                    then=Greatest(
                        "last_hospitalization_to",
                        "last_psychiatric_appointment_date",
                        "last_home_visit_by_doctor_date",
                    )
                ),
                When(
                    last_hospitalization_from__isnull=True,
                    last_hospitalization_to__isnull=True,
                    then=Greatest(
                        "last_psychiatric_appointment_date",
                        "last_home_visit_by_doctor_date",
                    )
                ),
                default=None,
                output_field=DateField()
            )
        ).annotate(
            interval=ExpressionWrapper(
                F("max_examination_interval") * Value(timedelta(days=1)),
                output_field=DurationField()
            ),
            deadline=ExpressionWrapper(
                F("last_date") + F("interval"),
                output_field=DateField()
            ),
            # deadline o‘tib ketganmi
            is_overdue=Case(
                When(last_date__isnull=True, then=Value(True)),
                When(deadline__lt=Value(today, output_field=DateField()), then=Value(True)),
                default=Value(False),
                output_field=BooleanField()
            )
        )

        qs = super().get_queryset(request).filter(filter_q).annotate(
            total_neighborhood=Count("neighborhoods", distinct=True),
            total_patients=Count("neighborhoods__patients", distinct=True),
            total_aggressive_patients=Count(
                "neighborhoods__patients",
                filter=Q(neighborhoods__patients__in=patients.filter(is_aggressive=True)),
                distinct=True
            ),
            total_convicted_patients=Count(
                "neighborhoods__patients",
                filter=Q(neighborhoods__patients__in=patients.filter(is_convicted=True)),
                distinct=True
            ),
            total_abroad_long_term_patients=Count(
                "neighborhoods__patients",
                filter=Q(neighborhoods__patients__in=patients.filter(is_abroad_long_term=True)),
                distinct=True
            ),
            late_count=Count(
                "neighborhoods__patients",
                filter=Q(neighborhoods__patients__in=patients.filter(is_overdue=True)),
                distinct=True,
            ),
            on_time_count=Count(
                "neighborhoods__patients",
                filter=Q(neighborhoods__patients__in=patients.filter(is_overdue=False)),
                distinct=True,
            ),
            aggressive_late_count=Count(
                "neighborhoods__patients",
                filter=Q(neighborhoods__patients__in=patients.filter(is_aggressive=True, is_overdue=True)),
                distinct=True,
            ),
            aggressive_on_time_count=Count(
                "neighborhoods__patients",
                filter=Q(neighborhoods__patients__in=patients.filter(is_aggressive=True, is_overdue=False)),
                distinct=True,
            ),
        )
        totals = defaultdict(int)
        for row in qs:
            totals["total_neighborhood"] += row.total_neighborhood
            totals["total_patients"] += row.total_patients
            totals["total_aggressive_patients"] += row.total_aggressive_patients
            totals["total_convicted_patients"] += row.total_convicted_patients
            totals["total_abroad_long_term_patients"] += row.total_abroad_long_term_patients
            totals["on_time_count"] += row.on_time_count
            totals["late_count"] += row.late_count
            totals["aggressive_on_time_count"] += row.aggressive_on_time_count
            totals["aggressive_late_count"] += row.aggressive_late_count

        self.totals = dict(totals)
        return qs

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    @admin.display(ordering="total_patients", description=_("Total patient count"))
    def total_patients(self, obj):
        return obj.total_patients

    @admin.display(ordering="total_aggressive_patients", description=_("Total aggressive patient count"))
    def total_aggressive_patients(self, obj):
        return obj.total_aggressive_patients

    @admin.display(ordering="total_convicted_patients", description=_("Total convicted patient count"))
    def total_convicted_patients(self, obj):
        return obj.total_convicted_patients

    @admin.display(ordering="total_abroad_long_term_patients", description=_("Total abroad long term patient count"))
    def total_abroad_long_term_patients(self, obj):
        return obj.total_abroad_long_term_patients

    @admin.display(ordering="total_neighborhood", description=_("Total neighborhood count"))
    def total_neighborhood(self, obj):
        return obj.total_neighborhood

    @admin.display(ordering="on_time_count", description=_("Number of mentally ill people who have not missed their follow-up check-up"))
    def on_time_count(self, obj):
        return obj.on_time_count

    @admin.display(ordering="late_count", description=_("Number of mentally ill people who missed their follow-up check-up"))
    def late_count(self, obj):
        return obj.late_count

    @admin.display(ordering="aggressive_on_time_count", description=_("Number of aggressive mentally ill people who have not missed their follow-up check-up"))
    def aggressive_on_time_count(self, obj):
        return obj.aggressive_on_time_count

    @admin.display(ordering="aggressive_late_count", description=_("Number of aggressive mentally ill people who missed their follow-up check-up"))
    def aggressive_late_count(self, obj):
        return obj.aggressive_late_count

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("export-excel/", self.admin_site.admin_view(self.export_as_excel), name="monitoring_export_excel"),
        ]
        return custom_urls + urls

    def export_as_excel(self, request):
        qs = self.get_changelist_instance(request).get_queryset(request)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "tuman"

        ws.append(["№", "Ҳудуд", "Жами маҳаллалар сони", "Жами руҳий касаллар сони",
                   "Жами тажовузкор руҳий касаллар сони",
                   "Жами муқаддам судланган руҳий касаллар сони",
                   "Жами узоқ муддатга кетган руҳий касаллар сони",
                   "Кейинги текширувни ўтказиб юборган руҳий касаллар сони",
                   "Кейинги текширувни ўтказиб юбормаган руҳий касаллар сони",
                   "Кейинги текширувни ўтказиб юборган тажовузкор руҳий касаллар сони",
                   "Кейинги текширувни ўтказиб юбормаган тажовузкор руҳий касаллар сони"])

        for index, obj in enumerate(qs, start=1):
            ws.append([
                index,
                obj.name,
                obj.total_neighborhood,
                obj.total_patients,
                obj.total_aggressive_patients,
                obj.total_convicted_patients,
                obj.total_abroad_long_term_patients,
                obj.late_count,
                obj.on_time_count,
                obj.aggressive_late_count,
                obj.aggressive_on_time_count,
            ])
        ws.append(["", "Жами", qs.aggregate(total_neighborhoods=Sum("total_neighborhood"))["total_neighborhoods"] or 0,
                qs.aggregate(total_patient=Sum("total_patients"))["total_patient"] or 0,
                qs.aggregate(total_aggressive_patient=Sum("total_aggressive_patients"))["total_aggressive_patient"] or 0,
                qs.aggregate(total_convicted_patient=Sum("total_convicted_patients"))["total_convicted_patient"] or 0,
                qs.aggregate(total_abroad_long_term_patient=Sum("total_abroad_long_term_patients"))["total_abroad_long_term_patients"] or 0,
                qs.aggregate(total_late_count=Sum("late_count"))["total_late_count"] or 0,
                qs.aggregate(total_on_time_count=Sum("on_time_count"))["total_on_time_count"] or 0,
                qs.aggregate(total_aggressive_late_count=Sum("aggressive_late_count"))["total_aggressive_late_count"] or 0,
                qs.aggregate(total_aggressive_on_time_count=Sum("aggressive_on_time_count"))["total_aggressive_on_time_count"] or 0
                ])

        uniform_width = 20
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = uniform_width
            for cell in col:
                cell.alignment = Alignment(wrap_text=True)

        today_str = datetime.date.today().strftime("%Y-%m-%d")
        filename = f"monitoring-{today_str}.xlsx"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename={filename}'
        return response

    def detail_export_as_excel(self, obj_name, qs, totals):

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "mahalla"

        ws.append(["№", "Маҳалла", "Жами руҳий касаллар сони",
                   "Жами тажовузкор руҳий касаллар сони",
                   "Жами муқаддам судланган руҳий касаллар сони",
                   "Жами узоқ муддатга кетган руҳий касаллар сони",
                   "Кейинги текширувни ўтказиб юборган руҳий касаллар сони",
                   "Кейинги текширувни ўтказиб юбормаган руҳий касаллар сони",
                   "Кейинги текширувни ўтказиб юборган тажовузкор руҳий касаллар сони",
                   "Кейинги текширувни ўтказиб юбормаган тажовузкор руҳий касаллар сони"])

        for index, obj in enumerate(qs, start=1):
            ws.append([
                index,
                obj.name,
                obj.total_patients,
                obj.total_aggressive_patients,
                obj.total_convicted_patients,
                obj.total_abroad_long_term_patients,
                obj.late_count,
                obj.on_time_count,
                obj.aggressive_late_count,
                obj.aggressive_on_time_count,
            ])
        ws.append(["", "Жами", totals["total_patients"], totals["total_aggressive_patients"],
                   totals["total_convicted_patients"], totals["total_abroad_long_term_patients"], totals["late_count"],
                   totals["on_time_count"], totals["aggressive_late_count"], totals["aggressive_on_time_count"]
                ])

        uniform_width = 20
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = uniform_width
            for cell in col:
                cell.alignment = Alignment(wrap_text=True)

        today_str = datetime.date.today().strftime("%Y-%m-%d")
        filename = f"monitoring-{today_str}.xlsx"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename={filename}'
        return response