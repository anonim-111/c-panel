import json
from datetime import date, datetime

from django.contrib.auth.models import Group
from django.core.management.base import BaseCommand
from django.db import transaction

from psytracks.models import Doctor, Patient, Psychiatrist, ReasonForSpecialConsideration, \
    ReceivingSupportiveTherapyChoices, SocialDomesticEnvironment, AlcoholAndDrugUse, WhereIsNow
from users.models import User
from utils.models import Neighborhood, District
from openpyxl import load_workbook

def get_birth_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    elif isinstance(value, str):
        try:
            return datetime.strptime(value.strip(), "%d.%m.%Y").date()
        except Exception as e:
            print(e, value)
            return None

def get_receiving_supportive_therapy(value):
    if value is None:
        return ""
    if value == "Мунтазам олаяпти" or value == "Мунтазам оляпти":
        return ReceivingSupportiveTherapyChoices.REGULARLY_RECEIVING
    elif value == "Камдан кам олаяпти":
        return ReceivingSupportiveTherapyChoices.RERALY_RECEIVING
    elif value == "Олмаяпти":
        return ReceivingSupportiveTherapyChoices.NOT_RECEIVING
    else:
        return ReceivingSupportiveTherapyChoices.QUICKLY_RECEIVING

def get_alcohol_and_drug_use(value):
    if value is None:
        return ""
    if value == "Истеъмол қилмайди":
        return AlcoholAndDrugUse.NOT_CONSUME
    else:
        return AlcoholAndDrugUse.AlCOHOL

def get_where_is_now(value):
    if value is None:
        return ""
    if value == "Уйда":
        return WhereIsNow.AT_HOME
    elif value == "Шифохонада":
        return WhereIsNow.IN_HOSPITAL
    elif value == "Ҳудудидан чиқиб кетган":
        return WhereIsNow.OUT_OF_THE_AREA
    else:
        return WhereIsNow.ADDRESS_UNKNOWN


class Command(BaseCommand):
    help = "Load district and neighborhood names from JSON file"

    def handle(self, *args, **options):

        wb = load_workbook("ruhiy-kasallar.xlsx")
        ws = wb.active

        with transaction.atomic():
            group = Group.objects.get(id=7)
            count = 4
            for row in ws.iter_rows(min_row=4, values_only=False):
                full_name = row[1].value
                print(full_name)
                if full_name is None:
                    break
                district = District.objects.filter(name=row[4].value.strip()).first()
                neighborhood = Neighborhood.objects.filter(district=district, name=row[5].value.strip()).first()
                inspector = neighborhood.inspector if hasattr(neighborhood, "inspector") else None
                if inspector is None:
                    print(district, neighborhood)
                psychiatrist = Psychiatrist.objects.filter(district=district, full_name=row[18].value.strip()).first()
                if psychiatrist is None:
                    user = User.objects.create(username=f"psixiatr{count}", is_staff=True)
                    user.groups.add(group)
                    user.set_password('1q2w!Q@W')
                    user.save()
                    count += 1
                    psychiatrist = Psychiatrist.objects.create(district=district, user=user, full_name=row[18].value.strip())

                social_domestic_environment_name = row[12].value
                if social_domestic_environment_name is None:
                    social_domestic_environment = None
                else:
                    social_domestic_environment, _ = SocialDomesticEnvironment.objects.get_or_create(name=social_domestic_environment_name)
                reason_for_special_consideration_name = row[7].value
                if reason_for_special_consideration_name is None:
                    reason_for_special_consideration = None
                else:
                    reason_for_special_consideration, _ = ReasonForSpecialConsideration.objects.get_or_create(name=reason_for_special_consideration_name)



                Patient.objects.create(
                    full_name=full_name,
                    pinfl=row[2].value,
                    birth_date=get_birth_date(row[3].value),
                    address=row[6].value,
                    last_psychiatric_appointment_date=get_birth_date(row[8].value),
                    last_home_visit_by_doctor_date=get_birth_date(row[9].value),
                    last_hospitalization_from=get_birth_date(row[14].value),
                    last_hospitalization_to=get_birth_date(row[15].value),
                    reason=row[10].value,
                    description_where_is_now=row[17].value,
                    reason_for_special_consideration=reason_for_special_consideration,
                    social_domestic_environment=social_domestic_environment,
                    receiving_supportive_therapy=get_receiving_supportive_therapy(str(row[11].value).strip()),
                    alcohol_and_drug_use=get_alcohol_and_drug_use(str(row[13].value).strip()),
                    where_is_now=get_where_is_now(str(row[16].value).strip()),
                    neighborhood=neighborhood,
                    psychiatrist=psychiatrist,
                    inspector=inspector,
                    is_aggressive=True
                )


            raise AssertionError


        self.stdout.write(self.style.SUCCESS("District and neighborhood data loaded successfully!"))
