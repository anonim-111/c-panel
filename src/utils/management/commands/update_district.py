import json
from datetime import date, datetime

from django.contrib.auth.models import Group
from django.core.management.base import BaseCommand
from django.db import transaction

from psytracks.models import Doctor
from users.models import User
from utils.models import Neighborhood
from openpyxl import load_workbook

def get_birth_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    elif isinstance(value, str):
        try:
            return datetime.strptime(value.strip(), "%d.%m.%Y").date()
        except Exception as e:
            print(e, value)
            return None

def get_phone(value):
    if value is None:
        return ""
    return value.replace(" ", "").replace(".", "").replace(",", "").replace("-", "")

class Command(BaseCommand):
    help = "Load district and neighborhood names from JSON file"

    def handle(self, *args, **options):

        # group = Group.objects.get(id=2)
        #
        # for i in range(64, 720):
        #     user = User.objects.create(
        #         username=f"uchaskavoy{i}",
        #         is_staff=True
        #     )
        #     user.set_password("1q2w!Q@W")
        #     user.groups.add(group)
        #     user.save()

        # neighborhoods = Neighborhood.objects.filter(inspector__isnull=True)
        # print(neighborhoods.count())
        # for neighborhood in neighborhoods:
        #     print(neighborhood.__str__())


        wb = load_workbook("doctors.xlsx")
        ws = wb.active

        # Merged hujayralarni tekshirish uchun
        merged_cells = ws.merged_cells.ranges

        with transaction.atomic():
            for row in ws.iter_rows(min_row=1, values_only=False):
                full_name = row[1].value
                if full_name is None:
                    break
                value = row[4].value
                if value is None:
                    for merged in merged_cells:
                        if row[4].coordinate in merged:
                            top_left = ws[merged.coord.split(":")[0]]
                            value = top_left.value
                            break
                Doctor.objects.create(
                    full_name=full_name,
                    birth_date=get_birth_date(row[2].value),
                    phone=get_phone(str(row[3].value)),
                    brigade_number=value,
                    polyclinic_name=row[5].value,
                    neighborhood=Neighborhood.objects.filter(district_id=3, name=row[6].value).first()
                )



            # raise AssertionError()

        self.stdout.write(self.style.SUCCESS("District and neighborhood data loaded successfully!"))
